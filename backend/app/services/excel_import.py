from typing import List, Tuple, Dict, Any
import json
from openpyxl import load_workbook
from io import BytesIO

REQUIRED_COLUMNS = [
    "title",
    "description",
    "complexity",
    "type",
    "options",
    "correct_answers",
    "max_score",
    "tags",
]

def parse_questions_excel(raw_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(filename=BytesIO(raw_bytes))
    ws = wb.active

    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value).strip() if cell.value else ""
        headers[header] = idx

    errors: List[Dict[str, Any]] = []
    valid_rows: List[Dict[str, Any]] = []

    for col in REQUIRED_COLUMNS:
        if col not in headers:
            errors.append({"row": 1, "error": f"Missing required column: {col}"})
            return [], errors

    for row_idx in range(2, ws.max_row + 1):
        row_data: Dict[str, Any] = {}
        try:
            title = ws.cell(row=row_idx, column=headers["title"]).value
            if not title:
                raise ValueError("title is required")
            row_data["title"] = str(title)

            row_data["description"] = ws.cell(row=row_idx, column=headers["description"]).value or ""
            row_data["complexity"] = ws.cell(row=row_idx, column=headers["complexity"]).value or ""

            q_type = ws.cell(row=row_idx, column=headers["type"]).value
            if q_type not in {"single_choice", "multi_choice", "text", "image_upload"}:
                raise ValueError("Invalid type")
            row_data["type"] = q_type

            options_raw = ws.cell(row=row_idx, column=headers["options"]).value or "[]"
            try:
                options = json.loads(options_raw)
            except Exception:
                raise ValueError("options must be valid JSON array")
            row_data["options"] = options

            correct_raw = ws.cell(row=row_idx, column=headers["correct_answers"]).value or "null"
            try:
                correct = json.loads(correct_raw)
            except Exception:
                raise ValueError("correct_answers must be valid JSON")
            row_data["correct_answers"] = correct

            max_score_val = ws.cell(row=row_idx, column=headers["max_score"]).value
            row_data["max_score"] = int(max_score_val) if max_score_val is not None else 1

            tags_raw = ws.cell(row=row_idx, column=headers["tags"]).value or ""
            tags = [t.strip() for t in str(tags_raw).split(",") if t.strip()]
            row_data["tags"] = tags

            valid_rows.append(row_data)
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    return valid_rows, errors
